"""
文档上传API（仅验证和保存）
"""
from fastapi import APIRouter, UploadFile, File, Query, HTTPException, Depends, Body
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import List, Optional, Dict, Any
import os
import uuid
import logging
import json
from datetime import datetime

from app.core.mysql_client import get_db
from app.models.document_upload import DocumentUpload, DocumentStatus
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/document-upload", tags=["文档上传"])


class DocumentUploadResponse(BaseModel):
    """文档上传响应"""
    id: int
    file_name: str
    file_size: int
    file_path: str
    file_extension: str
    status: str
    upload_time: str
    message: str


class DocumentListResponse(BaseModel):
    """文档列表响应"""
    documents: List[dict]
    total: int
    page: int
    page_size: int


@router.post("/upload", response_model=DocumentUploadResponse)
async def upload_document(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    上传文档（仅验证和保存，不进行解析处理）
    """
    try:
        # 验证文件格式（暂时只支持 .docx 格式）
        file_extension = os.path.splitext(file.filename)[1].lower()
        if file_extension != ".docx":
            raise HTTPException(status_code=400, detail="只支持 .docx 格式，暂不支持 .doc 格式")
        
        # 读取文件内容
        content = await file.read()
        file_size = len(content)
        
        # 验证文件大小（50MB限制）
        max_size = 50 * 1024 * 1024
        if file_size > max_size:
            raise HTTPException(status_code=400, detail=f"文件大小超过限制：最大支持 50MB，当前文件 {file_size / 1024 / 1024:.2f}MB")
        
        # 保存文件到uploads目录
        upload_dir = "uploads"
        os.makedirs(upload_dir, exist_ok=True)
        
        # 生成唯一文件名
        file_id = str(uuid.uuid4())
        file_name = file.filename or f"文档_{file_id}{file_extension}"
        file_path = os.path.join(upload_dir, f"{file_id}{file_extension}")
        
        # 保存文件
        with open(file_path, "wb") as f:
            f.write(content)
        
        logger.info(f"文件已保存: {file_path}, 大小: {file_size} 字节")
        
        # 保存到数据库
        db_document = DocumentUpload(
            file_name=file_name,
            file_size=file_size,
            file_path=file_path,
            file_extension=file_extension,
            status=DocumentStatus.VALIDATED
        )
        db.add(db_document)
        db.commit()
        db.refresh(db_document)
        
        logger.info(f"文档上传记录已保存到数据库: ID={db_document.id}")
        
        return DocumentUploadResponse(
            id=db_document.id,
            file_name=db_document.file_name,
            file_size=db_document.file_size,
            file_path=db_document.file_path,
            file_extension=db_document.file_extension,
            status=db_document.status.value,
            upload_time=db_document.upload_time.isoformat(),
            message="文件上传并验证成功"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"上传文档失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"上传失败: {str(e)}")


@router.get("/list", response_model=DocumentListResponse)
async def list_documents(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(10, ge=1, le=100, description="每页数量"),
    search: Optional[str] = Query(None, description="搜索关键词（文件名）"),
    status: Optional[str] = Query(None, description="状态筛选"),
    db: Session = Depends(get_db)
):
    """
    获取文档列表（支持分页、搜索、筛选）
    """
    try:
        # 构建查询
        query = db.query(DocumentUpload)
        
        # 搜索条件
        if search:
            query = query.filter(
                or_(
                    DocumentUpload.file_name.like(f"%{search}%"),
                    DocumentUpload.file_path.like(f"%{search}%")
                )
            )
        
        # 状态筛选
        if status:
            try:
                status_enum = DocumentStatus(status)
                query = query.filter(DocumentUpload.status == status_enum)
            except ValueError:
                pass  # 无效的状态值，忽略
        
        # 获取总数
        total = query.count()
        
        # 分页
        offset = (page - 1) * page_size
        documents = query.order_by(DocumentUpload.upload_time.desc()).offset(offset).limit(page_size).all()
        
        # 转换为字典列表
        documents_list = [doc.to_dict() for doc in documents]
        
        return DocumentListResponse(
            documents=documents_list,
            total=total,
            page=page,
            page_size=page_size
        )
        
    except Exception as e:
        logger.error(f"获取文档列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@router.get("/group-ids", response_model=List[str])
async def list_group_ids(
    search: Optional[str] = Query(None, description="搜索关键词（模糊匹配 Group ID）"),
    limit: int = Query(50, ge=1, le=200, description="返回数量限制"),
    db: Session = Depends(get_db)
):
    """
    获取所有已有的 Group ID 列表（用于下拉选择）
    支持模糊匹配搜索
    """
    try:
        # 查询所有非空的 document_id
        query = db.query(DocumentUpload.document_id).filter(
            DocumentUpload.document_id.isnot(None),
            DocumentUpload.document_id != ""
        )
        
        # 模糊匹配搜索
        if search:
            query = query.filter(DocumentUpload.document_id.contains(search))
        
        # 去重并排序
        group_ids = query.distinct().order_by(DocumentUpload.document_id).limit(limit).all()
        
        # 提取字符串列表
        result = [group_id[0] for group_id in group_ids if group_id[0]]
        
        return result
        
    except Exception as e:
        logger.error(f"获取 Group ID 列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@router.delete("/{document_id}")
async def delete_document(
    document_id: int,
    db: Session = Depends(get_db)
):
    """
    删除文档（删除数据库记录和文件）
    """
    try:
        # 查询文档
        document = db.query(DocumentUpload).filter(DocumentUpload.id == document_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 删除文件（处理相对路径）
        file_path_to_delete = document.file_path
        if not os.path.isabs(file_path_to_delete):
            file_path_to_delete = os.path.join("/app", file_path_to_delete)
        
        if os.path.exists(file_path_to_delete):
            try:
                os.remove(file_path_to_delete)
                logger.info(f"文件已删除: {file_path_to_delete}")
            except Exception as e:
                logger.warning(f"删除文件失败: {e}")
        
        # 删除Markdown文件
        if document.parsed_content_path:
            parsed_content_file_abs = os.path.join("/app", document.parsed_content_path)
            if os.path.exists(parsed_content_file_abs):
                try:
                    os.remove(parsed_content_file_abs)
                    logger.info(f"已删除parsed_content.md文件: {parsed_content_file_abs}")
                except Exception as e:
                    logger.warning(f"删除parsed_content.md文件失败: {e}")
        
        if document.summary_content_path:
            summary_content_file_abs = os.path.join("/app", document.summary_content_path)
            if os.path.exists(summary_content_file_abs):
                try:
                    os.remove(summary_content_file_abs)
                    logger.info(f"已删除summary_content.md文件: {summary_content_file_abs}")
                except Exception as e:
                    logger.warning(f"删除summary_content.md文件失败: {e}")
        
        if document.structured_content_path:
            structured_content_file_abs = os.path.join("/app", document.structured_content_path)
            if os.path.exists(structured_content_file_abs):
                try:
                    os.remove(structured_content_file_abs)
                    logger.info(f"已删除structured_content.json文件: {structured_content_file_abs}")
                except Exception as e:
                    logger.warning(f"删除structured_content.json文件失败: {e}")
        
        # 删除解析文件目录（如果为空）
        document_id_for_content = f"upload_{document_id}"
        parsed_content_dir = os.path.join("/app", "uploads", "parsed_content", document_id_for_content)
        if os.path.exists(parsed_content_dir):
            try:
                # 检查目录是否为空
                if not os.listdir(parsed_content_dir):
                    os.rmdir(parsed_content_dir)
                    logger.info(f"已删除空目录: {parsed_content_dir}")
            except Exception as e:
                logger.warning(f"删除目录失败: {e}")
        
        # 删除数据库记录
        db.delete(document)
        db.commit()
        
        return {"message": "文档删除成功", "id": document_id}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除文档失败: {e}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


@router.get("/{document_id}")
async def get_document(
    document_id: int,
    db: Session = Depends(get_db)
):
    """
    获取文档详情
    """
    try:
        document = db.query(DocumentUpload).filter(DocumentUpload.id == document_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        return document.to_dict()
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取文档详情失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@router.get("/{upload_id}/parsed-content")
async def get_parsed_content(
    upload_id: int,
    db: Session = Depends(get_db)
):
    """
    获取解析后的完整对应文档（parsed_content.md）
    """
    try:
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 检查文件路径是否存在（不依赖状态，只检查文件）
        if not document.parsed_content_path:
            raise HTTPException(status_code=404, detail="解析结果不存在，请先执行步骤2解析")
        
        # 读取Markdown文件
        parsed_content_file_abs = os.path.join("/app", document.parsed_content_path)
        if not os.path.exists(parsed_content_file_abs):
            raise HTTPException(status_code=404, detail="解析结果文件不存在，请重新执行步骤2解析")
        
        with open(parsed_content_file_abs, 'r', encoding='utf-8') as f:
            content = f.read()
        
        return {
            "upload_id": upload_id,
            "file_name": document.file_name,
            "content": content,
            "content_length": len(content)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取parsed_content失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@router.get("/{upload_id}/summary-content")
async def get_summary_content(
    upload_id: int,
    db: Session = Depends(get_db)
):
    """
    获取解析后的总结文档（summary_content.md）
    """
    try:
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 检查文件路径是否存在（不依赖状态，只检查文件）
        if not document.summary_content_path:
            raise HTTPException(status_code=404, detail="总结文档不存在，请先执行步骤2解析")
        
        # 读取Markdown文件
        summary_content_file_abs = os.path.join("/app", document.summary_content_path)
        if not os.path.exists(summary_content_file_abs):
            raise HTTPException(status_code=404, detail="总结文档文件不存在，请重新执行步骤2解析")
        
        with open(summary_content_file_abs, 'r', encoding='utf-8') as f:
            content = f.read()
        
        return {
            "upload_id": upload_id,
            "file_name": document.file_name,
            "content": content,
            "content_length": len(content)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取summary_content失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@router.get("/{upload_id}/structured-content")
async def get_structured_content(
    upload_id: int,
    db: Session = Depends(get_db)
):
    """
    获取解析后的结构化数据（structured_content.json）
    """
    try:
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 检查文件路径是否存在（不依赖状态，只检查文件）
        if not document.structured_content_path:
            raise HTTPException(status_code=404, detail="结构化数据不存在，请先执行步骤2解析")
        
        # 读取JSON文件
        structured_content_file_abs = os.path.join("/app", document.structured_content_path)
        if not os.path.exists(structured_content_file_abs):
            raise HTTPException(status_code=404, detail="结构化数据文件不存在，请重新执行步骤2解析")
        
        with open(structured_content_file_abs, 'r', encoding='utf-8') as f:
            content = json.load(f)
        
        return {
            "upload_id": upload_id,
            "file_name": document.file_name,
            "content": content,
            "item_count": len(content)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取structured_content失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@router.get("/{upload_id}/chunks")
async def get_chunks(
    upload_id: int,
    db: Session = Depends(get_db)
):
    """
    获取分块结果（chunks.json）
    """
    try:
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 检查文件路径是否存在（不依赖状态，只检查文件）
        if not document.chunks_path:
            raise HTTPException(status_code=404, detail="分块结果不存在，请先执行步骤4分块")
        
        # 读取 JSON 文件
        chunks_file_abs = os.path.join("/app", document.chunks_path)
        if not os.path.exists(chunks_file_abs):
            raise HTTPException(status_code=404, detail="分块结果文件不存在，请重新执行步骤4分块")
        
        with open(chunks_file_abs, 'r', encoding='utf-8') as f:
            content = json.load(f)
        
        return {
            "upload_id": upload_id,
            "file_name": document.file_name,
            "content": content
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取chunks失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


class DocumentParseResponse(BaseModel):
    """文档解析响应"""
    upload_id: int
    file_name: str
    document_id: str  # 用于构建图片和嵌入文档的URL（使用upload_id作为标识）
    parsed_content: str  # 完整对应文档（1:1对应原始文档）
    summary_content: str  # 总结文档（包含图片清单、表格清单等）
    structured_content: List[dict] = Field(default_factory=list)  # 结构化数据
    sections: List[dict] = Field(default_factory=list)
    statistics: dict = Field(default_factory=dict)


@router.post("/{upload_id}/parse", response_model=DocumentParseResponse)
async def parse_document(
    upload_id: int,
    max_tokens_per_section: int = Query(8000, ge=1000, le=20000, description="每个章节的最大token数"),
    db: Session = Depends(get_db)
):
    """
    解析文档（仅解析，不保存到Neo4j）
    """
    try:
        from app.services.word_document_service import WordDocumentService
        
        # 查询文档
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 将相对路径转换为绝对路径
        if not os.path.isabs(document.file_path):
            # 如果是相对路径，转换为绝对路径（相对于/app目录）
            file_path_abs = os.path.join("/app", document.file_path)
        else:
            file_path_abs = document.file_path
        
        # 检查文件是否存在
        if not os.path.exists(file_path_abs):
            raise HTTPException(status_code=404, detail=f"文件不存在: {file_path_abs}")
        
        logger.info(f"开始解析文档: upload_id={upload_id}, file_path={file_path_abs}")
        
        # 更新状态为解析中
        document.status = DocumentStatus.PARSING
        db.commit()
        
        try:
            # 解析文档（使用绝对路径）
            # 使用upload_id作为document_id，用于保存图片和OLE对象
            document_id_for_content = f"upload_{upload_id}"
            doc_data = WordDocumentService._parse_word_document(file_path_abs, document_id_for_content)
            logger.info(f"文档解析完成: {len(doc_data['structured_content'])} 个元素")
            
            # 按章节分块
            sections = WordDocumentService._split_by_sections(
                doc_data["structured_content"],
                max_tokens=max_tokens_per_section
            )
            logger.info(f"文档分为 {len(sections)} 个章节")
            
            # 构建解析后的内容（1:1对应原始文档）
            # 方案A：先输出第一个一级标题之前的所有内容（封面页、表格等），然后按章节输出
            parsed_content = ""
            
            # 找到第一个一级标题的位置
            first_level1_heading_idx = None
            for idx, item in enumerate(doc_data.get("structured_content", [])):
                if item.get("type") == "heading" and item.get("level", 1) == 1:
                    first_level1_heading_idx = idx
                    break
            
            # 输出第一个一级标题之前的所有内容（封面页、表格等）
            if first_level1_heading_idx is not None and first_level1_heading_idx > 0:
                # 有第一个一级标题，输出之前的内容
                prefix_content = WordDocumentService._build_content_from_items(
                    doc_data.get("structured_content", [])[:first_level1_heading_idx],
                    doc_data,
                    document_id_for_content,
                    upload_id
                )
                if prefix_content:
                    parsed_content += prefix_content + "\n\n"
            elif first_level1_heading_idx is None:
                # 没有一级标题，输出整个文档
                prefix_content = WordDocumentService._build_content_from_items(
                    doc_data.get("structured_content", []),
                    doc_data,
                    document_id_for_content,
                    upload_id
                )
                if prefix_content:
                    parsed_content += prefix_content + "\n\n"
            
            # 按章节输出每个章节的内容
            section_list = []
            for idx, section in enumerate(sections):
                section_content = WordDocumentService._build_section_content(
                    section, doc_data, idx, document_id_for_content, upload_id
                )
                if section_content:  # 只添加非空内容
                    parsed_content += section_content + "\n\n"
                
                # 构建章节信息
                section_list.append({
                    "title": section.get("title", ""),
                    "content": section_content[:500] + "..." if len(section_content) > 500 else section_content,
                    "level": section.get("level", 1),
                    "token_count": section.get("token_count", 0),
                    "section_id": section.get("section_id", f"section_{idx}")
                })
            
            # 构建总结文档
            summary_content = WordDocumentService._build_summary_content(
                doc_data, sections, document_id_for_content, upload_id, document.file_name
            )
            
            # 保存Markdown文件到文件系统
            parsed_content_dir = os.path.join("/app", "uploads", "parsed_content", document_id_for_content)
            os.makedirs(parsed_content_dir, exist_ok=True)
            
            parsed_content_file = os.path.join(parsed_content_dir, "parsed_content.md")
            summary_content_file = os.path.join(parsed_content_dir, "summary_content.md")
            
            # 写入文件（覆盖模式）
            with open(parsed_content_file, 'w', encoding='utf-8') as f:
                f.write(parsed_content.strip())
            
            with open(summary_content_file, 'w', encoding='utf-8') as f:
                f.write(summary_content.strip())
            
            # 保存structured_content.json（结构化数据）
            structured_content_file = os.path.join(parsed_content_dir, "structured_content.json")
            with open(structured_content_file, 'w', encoding='utf-8') as f:
                json.dump(doc_data.get("structured_content", []), f, ensure_ascii=False, indent=2)
            
            # 保存文件路径到数据库（相对路径）
            parsed_content_path = f"uploads/parsed_content/{document_id_for_content}/parsed_content.md"
            summary_content_path = f"uploads/parsed_content/{document_id_for_content}/summary_content.md"
            structured_content_path = f"uploads/parsed_content/{document_id_for_content}/structured_content.json"
            
            document.parsed_content_path = parsed_content_path
            document.summary_content_path = summary_content_path
            document.structured_content_path = structured_content_path
            
            # 更新状态为已解析
            document.status = DocumentStatus.PARSED
            db.commit()
            
            logger.info(f"解析结果文件已保存: {parsed_content_path}, {summary_content_path}, {structured_content_path}")
            
            # 构建返回结果
            return DocumentParseResponse(
                upload_id=upload_id,
                file_name=document.file_name,
                document_id=document_id_for_content,  # 返回document_id，用于前端构建URL
                parsed_content=parsed_content.strip(),  # 完整对应文档（1:1对应原始文档）
                summary_content=summary_content.strip(),  # 总结文档
                structured_content=doc_data.get("structured_content", []),  # 结构化数据
                sections=section_list,
                statistics={
                    "total_sections": len(sections),
                    "total_images": len(doc_data.get("images", [])),
                    "total_tables": len(doc_data.get("tables", [])),
                    "total_links": len(doc_data.get("links", [])),
                    "total_ole_objects": len(doc_data.get("ole_objects", [])),
                    "text_length": len(doc_data.get("text_content", ""))
                }
            )
            
        except Exception as parse_error:
            # 解析失败，更新状态为错误
            document.status = DocumentStatus.ERROR
            document.error_message = str(parse_error)
            db.commit()
            logger.error(f"解析文档失败: {parse_error}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"解析失败: {str(parse_error)}")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"解析文档失败: {e}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")


class DocumentVersionResponse(BaseModel):
    """版本信息响应"""
    upload_id: int
    document_name: str
    base_name: str
    version: str
    version_number: int
    group_id: str
    date_str: str


class UpdateVersionRequest(BaseModel):
    """更新版本信息请求"""
    version: str
    version_number: int
    group_id: str


class SectionInfo(BaseModel):
    """章节信息"""
    section_id: str
    title: str
    level: int
    token_count: int
    content_length: int
    image_count: int
    table_count: int
    is_split: bool = False  # 是否被分割（超过max_tokens）


class ChunkInfo(BaseModel):
    """分块信息（用于保存到 chunks.json）"""
    chunk_id: str
    title: str
    level: int
    token_count: int
    start_index: int  # 在 structured_content 中的起始位置
    end_index: int    # 在 structured_content 中的结束位置
    content: str      # 该块的 Markdown 内容


class DocumentSplitResponse(BaseModel):
    """文档分块响应"""
    upload_id: int
    document_name: str
    strategy: str = "level_1"  # 分块策略
    max_tokens_per_section: int
    sections: List[SectionInfo] = Field(default_factory=list)
    statistics: Dict[str, Any] = Field(default_factory=dict)
    chunks_saved: bool = False  # 是否已保存 chunks.json


class DocumentProcessResponse(BaseModel):
    """文档处理响应"""
    upload_id: int
    document_id: str  # group_id
    document_name: str
    version: str
    version_number: int
    episodes: dict = Field(default_factory=dict)
    statistics: dict = Field(default_factory=dict)


class BuildCommunitiesRequest(BaseModel):
    """构建Community请求"""
    scope: str = Field(..., description="构建范围: 'current' 或 'cross'")
    group_ids: Optional[List[str]] = Field(None, description="跨文档时提供的group_id列表")


class CommunityInfo(BaseModel):
    """Community信息"""
    uuid: str
    name: str
    summary: str
    entity_count: int
    group_ids: List[str] = Field(default_factory=list)


class BuildCommunitiesResponse(BaseModel):
    """构建Community响应"""
    success: bool
    communities: List[CommunityInfo] = Field(default_factory=list)
    statistics: dict = Field(default_factory=dict)


@router.post("/{upload_id}/version", response_model=DocumentVersionResponse)
async def generate_version(
    upload_id: int,
    custom_group_id: Optional[str] = Query(None, description="用户自定义的 Group ID（可选）"),
    db: Session = Depends(get_db)
):
    """
    生成版本信息（提取版本号、生成group_id）
    
    步骤：
    1. 提取基础名称（去除版本号）
    2. 提取版本号
    3. 生成group_id（如果用户提供了自定义 Group ID，则使用用户的；否则自动生成）
    """
    try:
        from app.services.word_document_service import WordDocumentService
        from datetime import datetime
        import re
        
        # 查询文档
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 检查文档状态（必须是已解析状态）
        if document.status != DocumentStatus.PARSED:
            raise HTTPException(status_code=400, detail=f"文档状态不正确，需要已解析状态，当前状态：{document.status.value}")
        
        # 提取版本信息
        base_name = WordDocumentService._extract_base_name(document.file_name)
        version, version_number = WordDocumentService._extract_version(document.file_name)
        
        # 处理 Group ID
        if custom_group_id and custom_group_id.strip():
            # 用户提供了自定义 Group ID，验证格式
            group_id = custom_group_id.strip()
            # Graphiti 要求：只能包含字母数字、破折号、下划线
            if not re.match(r'^[a-zA-Z0-9\-_]+$', group_id):
                raise HTTPException(
                    status_code=400,
                    detail="Group ID 格式不正确，只能包含字母、数字、破折号(-)和下划线(_)"
                )
        else:
            # 自动生成 Group ID
            safe_base_name = WordDocumentService._sanitize_group_id(base_name)
            date_str = datetime.now().strftime('%Y%m%d')
            group_id = f"doc_{safe_base_name}_{date_str}"
        
        logger.info(f"生成版本信息: upload_id={upload_id}, base_name={base_name}, version={version} (版本号: {version_number}), group_id={group_id}")
        
        # 保存group_id到数据库的document_id字段
        document.document_id = group_id
        db.commit()
        
        return DocumentVersionResponse(
            upload_id=upload_id,
            document_name=document.file_name,
            base_name=base_name,
            version=version,
            version_number=version_number,
            group_id=group_id,
            date_str=datetime.now().strftime('%Y%m%d')
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成版本信息失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"生成版本信息失败: {str(e)}")


@router.put("/{upload_id}/version", response_model=DocumentVersionResponse)
async def update_version(
    upload_id: int,
    request: UpdateVersionRequest = Body(...),
    db: Session = Depends(get_db)
):
    """
    更新版本信息（版本号、group_id）
    """
    try:
        import re
        from datetime import datetime
        
        # 查询文档
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 验证 Group ID 格式
        if not re.match(r'^[a-zA-Z0-9\-_]+$', request.group_id):
            raise HTTPException(
                status_code=400,
                detail="Group ID 格式不正确，只能包含字母、数字、破折号(-)和下划线(_)"
            )
        
        # 验证版本号
        if request.version_number < 1:
            raise HTTPException(status_code=400, detail="版本号必须大于0")
        
        # 提取基础名称（用于返回）
        from app.services.word_document_service import WordDocumentService
        base_name = WordDocumentService._extract_base_name(document.file_name)
        
        logger.info(f"更新版本信息: upload_id={upload_id}, version={request.version}, version_number={request.version_number}, group_id={request.group_id}")
        
        # 保存group_id到数据库的document_id字段
        document.document_id = request.group_id
        db.commit()
        
        return DocumentVersionResponse(
            upload_id=upload_id,
            document_name=document.file_name,
            base_name=base_name,
            version=request.version,
            version_number=request.version_number,
            group_id=request.group_id,
            date_str=datetime.now().strftime('%Y%m%d')
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新版本信息失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"更新版本信息失败: {str(e)}")


@router.post("/{upload_id}/split", response_model=DocumentSplitResponse)
async def split_document(
    upload_id: int,
    strategy: str = Query("level_1", description="分块策略: auto(智能分块), level_1(一级标题), level_2(二级标题), level_3(三级标题), level_4(四级标题), level_5(五级标题), fixed_token(固定token), no_split(不分块)"),
    max_tokens_per_section: int = Query(8000, ge=1000, le=20000, description="每个章节的最大token数"),
    save_chunks: bool = Query(True, description="是否保存分块结果到 chunks.json"),
    db: Session = Depends(get_db)
):
    """
    章节分块（支持多种策略）
    
    策略：
    - auto: 智能分块（自动分析文档结构，选择最佳策略）
    - level_1: 按一级标题分块（默认）
    - level_2: 按二级标题分块
    - level_3: 按三级标题分块
    - level_4: 按四级标题分块
    - level_5: 按五级标题分块
    - fixed_token: 按固定token数分块
    - no_split: 不分块（整个文档作为一个块）
    
    步骤：
    1. 读取 structured_content.json
    2. 根据策略分块（如果是auto，先自动选择策略）
    3. 保存 chunks.json（可选）
    4. 返回分块结果
    """
    try:
        from app.services.word_document_service import WordDocumentService
        
        # 验证策略参数（支持 "auto" 智能分块）
        valid_strategies = ["auto", "level_1", "level_2", "level_3", "level_4", "level_5", "fixed_token", "no_split"]
        if strategy not in valid_strategies:
            raise HTTPException(status_code=400, detail=f"无效的分块策略: {strategy}，有效值: {valid_strategies}")
        
        # 查询文档
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 检查文档状态（允许已解析或已分块状态，支持重复分块覆盖）
        allowed_statuses = [DocumentStatus.PARSED, DocumentStatus.CHUNKED]
        if document.status not in allowed_statuses:
            raise HTTPException(status_code=400, detail=f"文档状态不正确，需要已解析或已分块状态，当前状态：{document.status.value}")
        
        logger.info(f"开始章节分块: upload_id={upload_id}, strategy={strategy}, max_tokens={max_tokens_per_section}")
        
        # 读取 structured_content.json 文件
        structured_content = None
        if document.structured_content_path:
            structured_content_file_abs = os.path.join("/app", document.structured_content_path)
            if os.path.exists(structured_content_file_abs):
                with open(structured_content_file_abs, 'r', encoding='utf-8') as f:
                    structured_content = json.load(f)
                logger.info(f"已读取structured_content.json: {len(structured_content)} 个元素")
        
        # 如果 structured_content.json 不存在，重新解析
        if structured_content is None:
            # 将相对路径转换为绝对路径
            if not os.path.isabs(document.file_path):
                file_path_abs = os.path.join("/app", document.file_path)
            else:
                file_path_abs = document.file_path
            
            if not os.path.exists(file_path_abs):
                raise HTTPException(status_code=404, detail=f"原始文件不存在: {file_path_abs}")
            
            document_id_for_content = f"upload_{upload_id}"
            doc_data = WordDocumentService._parse_word_document(file_path_abs, document_id_for_content)
            structured_content = doc_data.get("structured_content", [])
            logger.info(f"重新解析文档: {len(structured_content)} 个元素")
        
        # 如果策略是 "auto"，使用智能分块服务自动选择策略
        actual_strategy = strategy
        if strategy == "auto":
            from app.services.smart_chunking_service import SmartChunkingService
            import asyncio
            
            smart_chunking = SmartChunkingService()
            logger.info("使用智能分块服务自动选择最佳策略")
            
            # 创建事件循环（如果不存在）
            try:
                loop = asyncio.get_event_loop()
            except RuntimeError:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
            
            # 调用智能分块服务
            strategy_result = loop.run_until_complete(
                smart_chunking.analyze_and_select_strategy(
                    structured_content=structured_content,
                    document_name=document.file_name,
                    metadata=None
                )
            )
            actual_strategy = strategy_result["strategy"]
            logger.info(
                f"智能分块选择策略: {actual_strategy}, "
                f"原因: {strategy_result.get('reason', '')}, "
                f"置信度: {strategy_result.get('confidence', 0):.2f}"
            )
        
        # 根据策略进行分块
        sections = WordDocumentService._split_by_sections_with_strategy(
            structured_content,
            strategy=actual_strategy,
            max_tokens=max_tokens_per_section
        )
        logger.info(f"分块完成: {len(sections)} 个块，策略: {actual_strategy}（原始策略: {strategy}）")
        
        # 构建章节信息列表和 chunks 数据
        section_infos = []
        chunks_data = []
        total_tokens = 0
        max_tokens_count = 0
        min_tokens_count = float('inf')
        oversized_sections = []  # 超过max_tokens的章节（被分割的章节）
        
        for idx, section in enumerate(sections):
            token_count = section.get("token_count", 0)
            total_tokens += token_count
            max_tokens_count = max(max_tokens_count, token_count)
            min_tokens_count = min(min_tokens_count, token_count) if token_count > 0 else min_tokens_count
            
            # 检查是否被分割（标题包含"（续）"表示被分割）
            is_split = "（续）" in section.get("title", "")
            if is_split:
                oversized_sections.append(section.get("title", ""))
            
            section_infos.append(SectionInfo(
                section_id=section.get("section_id", f"chunk_{idx+1}"),
                title=section.get("title", ""),
                level=section.get("level", 1),
                token_count=token_count,
                content_length=len(section.get("content", "")),
                image_count=len(section.get("images", [])),
                table_count=len(section.get("tables", [])),
                is_split=is_split
            ))
            
            # 构建 chunk 数据（用于保存）
            chunks_data.append({
                "chunk_id": section.get("section_id", f"chunk_{idx+1}"),
                "title": section.get("title", ""),
                "level": section.get("level", 1),
                "token_count": token_count,
                "start_index": section.get("start_index", 0),
                "end_index": section.get("end_index", 0),
                "content": section.get("content", "")
            })
        
        # 计算统计信息
        avg_tokens = total_tokens / len(sections) if sections else 0
        statistics = {
            "total_sections": len(sections),
            "total_tokens": total_tokens,
            "avg_tokens": round(avg_tokens, 2),
            "max_tokens": max_tokens_count,
            "min_tokens": int(min_tokens_count) if min_tokens_count != float('inf') else 0,
            "oversized_sections_count": len(oversized_sections),
            "oversized_sections": oversized_sections
        }
        
        chunks_saved = False
        
        # 保存 chunks.json 文件（如果需要）
        if save_chunks:
            document_id_for_content = f"upload_{upload_id}"
            chunks_dir = os.path.join("/app", "uploads", "parsed_content", document_id_for_content)
            os.makedirs(chunks_dir, exist_ok=True)
            
            chunks_file = os.path.join(chunks_dir, "chunks.json")
            chunks_json = {
                "strategy": actual_strategy,  # 保存实际使用的策略（如果是auto，保存自动选择的策略）
                "original_strategy": strategy,  # 保存原始策略（auto或手动选择的策略）
                "max_tokens": max_tokens_per_section,
                "created_at": datetime.now().isoformat(),
                "total_chunks": len(chunks_data),
                "statistics": statistics,
                "chunks": chunks_data
            }
            
            with open(chunks_file, 'w', encoding='utf-8') as f:
                json.dump(chunks_json, f, ensure_ascii=False, indent=2)
            
            # 更新数据库
            chunks_path = f"uploads/parsed_content/{document_id_for_content}/chunks.json"
            document.chunks_path = chunks_path
            document.status = DocumentStatus.CHUNKED
            db.commit()
            
            chunks_saved = True
            logger.info(f"chunks.json 已保存: {chunks_path}")
        
        logger.info(f"章节分块完成: {len(sections)} 个块, 总token数: {total_tokens}, 策略: {strategy}")
        
        return DocumentSplitResponse(
            upload_id=upload_id,
            document_name=document.file_name,
            strategy=actual_strategy,  # 返回实际使用的策略
            max_tokens_per_section=max_tokens_per_section,
            sections=section_infos,
            statistics=statistics,
            chunks_saved=chunks_saved
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"章节分块失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"章节分块失败: {str(e)}")


@router.post("/{upload_id}/process", response_model=DocumentProcessResponse)
async def process_document(
    upload_id: int,
    provider: str = Query("qianwen", description="LLM提供商"),
    max_tokens_per_section: int = Query(8000, ge=1000, le=20000, description="每个章节的最大token数"),
    use_thinking: bool = Query(False, description="是否启用Thinking模式（仅本地大模型支持）"),
    db: Session = Depends(get_db)
):
    """
    处理文档并保存到Neo4j（方案B：混合使用parsed_content和summary_content）
    
    优化：复用步骤3和步骤4的结果
    - 如果步骤3已完成，使用数据库中的document_id（group_id）
    - 如果步骤4已完成，使用chunks.json的分块结果
    
    步骤：
    1. 版本管理（复用步骤3的group_id，或重新生成）
    2. 章节分块（复用步骤4的chunks.json，或重新分块）
    3. 创建Episode并保存到Neo4j：
       - 文档级Episode：使用summary_content的"文档概览"部分
       - 章节级Episode：使用chunks.json的content（如果存在）
       - 图片/表格Episode：关联到所属章节Episode
    """
    try:
        from app.services.word_document_service import WordDocumentService
        from app.core.graphiti_client import get_graphiti_instance
        from app.models.graphiti_entities import ENTITY_TYPES, EDGE_TYPES, EDGE_TYPE_MAP
        from app.core.neo4j_client import neo4j_client
        
        # 查询文档
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 检查文档状态（允许已解析、已分块、错误状态）
        # 错误状态的文档可以重新处理（会重新解析）
        allowed_statuses = [DocumentStatus.PARSED, DocumentStatus.CHUNKED, DocumentStatus.ERROR]
        if document.status not in allowed_statuses:
            raise HTTPException(status_code=400, detail=f"文档状态不正确，需要已解析、已分块或错误状态，当前状态：{document.status.value}")
        
        # 将相对路径转换为绝对路径
        if not os.path.isabs(document.file_path):
            file_path_abs = os.path.join("/app", document.file_path)
        else:
            file_path_abs = document.file_path
        
        if not os.path.exists(file_path_abs):
            raise HTTPException(status_code=404, detail=f"文件不存在: {file_path_abs}")
        
        logger.info(f"开始处理文档: upload_id={upload_id}, file_path={file_path_abs}")
        
        # 更新状态为处理中
        document.status = DocumentStatus.CHUNKING
        db.commit()
        
        try:
            # 步骤1：复用步骤3的group_id（如果存在）
            group_id = None
            version = None
            version_number = None
            
            if document.document_id:
                # 步骤3已完成，复用group_id
                group_id = document.document_id
                # 从文件名提取版本信息（用于Episode的版本字段）
                base_name = WordDocumentService._extract_base_name(document.file_name)
                version, version_number = WordDocumentService._extract_version(document.file_name)
                logger.info(f"复用步骤3的group_id: {group_id}, version={version}, version_number={version_number}")
            else:
                # 步骤3未完成，需要生成group_id
                logger.warning("步骤3未完成，将自动生成group_id")
                base_name = WordDocumentService._extract_base_name(document.file_name)
                version, version_number = WordDocumentService._extract_version(document.file_name)
                safe_base_name = WordDocumentService._sanitize_group_id(base_name)
                from datetime import datetime
                date_str = datetime.now().strftime('%Y%m%d')
                group_id = f"doc_{safe_base_name}_{date_str}"
                # 保存到数据库
                document.document_id = group_id
                db.commit()
                logger.info(f"自动生成group_id: {group_id}")
            
            # 读取Markdown文件（如果存在）
            parsed_content = None
            summary_content = None
            
            if document.parsed_content_path:
                parsed_content_file_abs = os.path.join("/app", document.parsed_content_path)
                if os.path.exists(parsed_content_file_abs):
                    with open(parsed_content_file_abs, 'r', encoding='utf-8') as f:
                        parsed_content = f.read()
                    logger.info(f"已读取parsed_content.md文件: {document.parsed_content_path}")
                else:
                    logger.warning(f"parsed_content.md文件不存在: {parsed_content_file_abs}，将重新解析")
            
            if document.summary_content_path:
                summary_content_file_abs = os.path.join("/app", document.summary_content_path)
                if os.path.exists(summary_content_file_abs):
                    with open(summary_content_file_abs, 'r', encoding='utf-8') as f:
                        summary_content = f.read()
                    logger.info(f"已读取summary_content.md文件: {document.summary_content_path}")
                else:
                    logger.warning(f"summary_content.md文件不存在: {summary_content_file_abs}，将重新解析")
            
            # 读取chunks.json（如果步骤4已完成）
            chunks_data = None
            if document.chunks_path:
                chunks_file_abs = os.path.join("/app", document.chunks_path)
                if os.path.exists(chunks_file_abs):
                    with open(chunks_file_abs, 'r', encoding='utf-8') as f:
                        chunks_data = json.load(f)
                    logger.info(f"已读取chunks.json: {document.chunks_path}, 包含 {len(chunks_data.get('chunks', []))} 个chunks")
                else:
                    logger.warning(f"chunks.json文件不存在: {chunks_file_abs}，将重新分块")
            
            # 如果文件不存在，需要重新解析文档
            document_id_for_content = f"upload_{upload_id}"
            doc_data = None
            if parsed_content is None or summary_content is None:
                doc_data = WordDocumentService._parse_word_document(file_path_abs, document_id_for_content)
                logger.info(f"文档解析完成: {len(doc_data['structured_content'])} 个元素")
                
                # 重新构建parsed_content和summary_content
                if parsed_content is None:
                    # 构建parsed_content（1:1对应原始文档）
                    parsed_content = ""
                    first_level1_heading_idx = None
                    for idx, item in enumerate(doc_data.get("structured_content", [])):
                        if item.get("type") == "heading" and item.get("level", 1) == 1:
                            first_level1_heading_idx = idx
                            break
                    
                    if first_level1_heading_idx is not None and first_level1_heading_idx > 0:
                        prefix_content = WordDocumentService._build_content_from_items(
                            doc_data.get("structured_content", [])[:first_level1_heading_idx],
                            doc_data,
                            document_id_for_content,
                            upload_id
                        )
                        if prefix_content:
                            parsed_content += prefix_content + "\n\n"
                    elif first_level1_heading_idx is None:
                        prefix_content = WordDocumentService._build_content_from_items(
                            doc_data.get("structured_content", []),
                            doc_data,
                            document_id_for_content,
                            upload_id
                        )
                        if prefix_content:
                            parsed_content += prefix_content + "\n\n"
                    
                    # 如果chunks.json不存在，需要重新分块
                    if chunks_data is None:
                        sections = WordDocumentService._split_by_sections(
                            doc_data["structured_content"],
                            max_tokens=max_tokens_per_section
                        )
                        logger.info(f"文档分为 {len(sections)} 个章节（重新分块）")
                        
                        # 按章节输出每个章节的内容
                        for idx, section in enumerate(sections):
                            section_content = WordDocumentService._build_section_content(
                                section, doc_data, idx, document_id_for_content, upload_id
                            )
                            if section_content:
                                parsed_content += section_content + "\n\n"
                
                if summary_content is None:
                    # 如果chunks.json不存在，需要重新分块
                    if chunks_data is None:
                        sections = WordDocumentService._split_by_sections(
                            doc_data["structured_content"],
                            max_tokens=max_tokens_per_section
                        )
                    else:
                        # 使用chunks.json中的sections信息（用于构建summary_content）
                        sections = []
                        for chunk in chunks_data.get('chunks', []):
                            sections.append({
                                'title': chunk.get('title', ''),
                                'level': chunk.get('level', 1)
                            })
                    
                    # 构建summary_content
                    summary_content = WordDocumentService._build_summary_content(
                        doc_data, sections, document_id_for_content, upload_id, document.file_name
                    )
            
            # 如果doc_data不存在，需要解析来获取structured_content（用于图片/表格Episode）
            if doc_data is None:
                doc_data = WordDocumentService._parse_word_document(file_path_abs, document_id_for_content)
                logger.info(f"文档解析完成: {len(doc_data['structured_content'])} 个元素（用于创建图片/表格Episode）")
            
            # 准备章节数据（用于创建章节级Episode）
            # 优先使用chunks.json，否则从parsed_content分割
            section_episodes_data = []
            if chunks_data and chunks_data.get('chunks'):
                # 使用chunks.json的content
                logger.info(f"使用chunks.json创建章节级Episode: {len(chunks_data['chunks'])} 个chunks")
                for chunk in chunks_data['chunks']:
                    section_episodes_data.append({
                        "chunk_id": chunk.get('chunk_id', ''),
                        "title": chunk.get('title', ''),
                        "level": chunk.get('level', 1),
                        "content": chunk.get('content', ''),
                        "start_index": chunk.get('start_index', 0),
                        "end_index": chunk.get('end_index', 0)
                    })
            elif parsed_content:
                # 从parsed_content按章节标题分割
                logger.info("从parsed_content分割章节（chunks.json不存在）")
                lines = parsed_content.split('\n')
                current_section = None
                current_content = []
                
                for line in lines:
                    # 检查是否是章节标题（一级标题，以#开头，后面跟空格）
                    if line.strip().startswith('# ') and not line.strip().startswith('##'):
                        # 保存上一个章节
                        if current_section:
                            section_episodes_data.append({
                                "chunk_id": f"chunk_{len(section_episodes_data) + 1}",
                                "title": current_section,
                                "level": 1,
                                "content": '\n'.join(current_content).strip(),
                                "start_index": 0,
                                "end_index": 0
                            })
                        # 开始新章节
                        current_section = line.strip()[2:].strip()  # 去掉"# "
                        current_content = [line]
                    else:
                        if current_section:
                            current_content.append(line)
                
                # 保存最后一个章节
                if current_section:
                    section_episodes_data.append({
                        "chunk_id": f"chunk_{len(section_episodes_data) + 1}",
                        "title": current_section,
                        "level": 1,
                        "content": '\n'.join(current_content).strip(),
                        "start_index": 0,
                        "end_index": 0
                    })
            
            # 步骤4：获取Graphiti实例
            graphiti = get_graphiti_instance(provider)
            
            # 步骤5：创建文档级Episode（使用summary_content的"文档概览"部分）
            # 从summary_content中提取"文档概览"部分
            summary_lines = summary_content.split('\n')
            overview_start = None
            overview_end = None
            for idx, line in enumerate(summary_lines):
                if line.strip() == "## 文档概览":
                    overview_start = idx + 1
                elif overview_start is not None and line.startswith("## ") and line.strip() != "## 文档概览":
                    overview_end = idx
                    break
            
            if overview_start is not None:
                overview_content = '\n'.join(summary_lines[overview_start:overview_end if overview_end else len(summary_lines)])
            else:
                # 如果没有找到"文档概览"，使用整个summary_content的前1000字符
                overview_content = summary_content[:1000]
            
            document_episode = await graphiti.add_episode(
                name=f"{document.file_name}_文档概览",
                episode_body=overview_content,
                source_description="Word文档",
                reference_time=doc_data["metadata"].get("created") or datetime.now(),
                entity_types={
                    "Requirement": ENTITY_TYPES.get("Requirement"),
                    "Document": ENTITY_TYPES.get("Document"),
                } if ENTITY_TYPES.get("Requirement") and ENTITY_TYPES.get("Document") else None,
                group_id=group_id
            )
            document_episode_uuid = document_episode.episode.uuid
            logger.info(f"文档级 Episode 创建完成: {document_episode_uuid}")
            
            # 更新文档级Episode的版本信息
            update_version_query = """
            MATCH (e:Episodic)
            WHERE e.uuid = $episode_uuid
            SET e.version = $version,
                e.version_number = $version_number,
                e.document_name = $document_name,
                e.file_path = $file_path,
                e.original_filename = $original_filename
            RETURN e.uuid as uuid
            """
            neo4j_client.execute_write(update_version_query, {
                "episode_uuid": document_episode_uuid,
                "version": version,
                "version_number": version_number,
                "document_name": document.file_name,
                "file_path": file_path_abs,
                "original_filename": os.path.basename(file_path_abs)
            })
            
            # 步骤6：创建章节级Episode（使用chunks.json的content或parsed_content）
            section_episodes = []
            section_episode_map = {}  # 用于图片/表格关联：{chunk_id: episode_uuid}
            
            for idx, section_data in enumerate(section_episodes_data):
                section_episode = await graphiti.add_episode(
                    name=f"{document.file_name}_章节_{idx+1}_{section_data['title'][:20]}",
                    episode_body=section_data["content"],
                    source_description="Word文档章节",
                    reference_time=doc_data["metadata"].get("created") or datetime.now(),
                    entity_types=ENTITY_TYPES,
                    edge_types=EDGE_TYPES,
                    edge_type_map=EDGE_TYPE_MAP,
                    group_id=group_id,
                    previous_episode_uuids=[document_episode_uuid]
                )
                # 更新版本信息
                neo4j_client.execute_write(update_version_query, {
                    "episode_uuid": section_episode.episode.uuid,
                    "version": version,
                    "version_number": version_number,
                    "document_name": document.file_name,
                    "file_path": file_path_abs,
                    "original_filename": os.path.basename(file_path_abs)
                })
                section_episode_uuid = section_episode.episode.uuid
                section_episodes.append(section_episode_uuid)
                
                # 建立chunk_id到episode_uuid的映射（用于图片/表格关联）
                chunk_id = section_data.get('chunk_id', f"chunk_{idx+1}")
                section_episode_map[chunk_id] = section_episode_uuid
                
                logger.info(f"章节级 Episode {idx+1} 创建完成: {section_episode_uuid} (chunk_id: {chunk_id})")
            
            # 辅助函数：根据图片/表格在structured_content中的位置，找到对应的chunk_id
            def find_chunk_for_item(item_position, structured_content, chunks_data):
                """
                根据item在structured_content中的位置，找到对应的chunk_id
                
                Args:
                    item_position: 图片/表格在structured_content中的索引位置
                    structured_content: 结构化内容数组
                    chunks_data: chunks.json数据
                
                Returns:
                    chunk_id或None
                """
                if not chunks_data or not chunks_data.get('chunks'):
                    return None
                
                # 遍历chunks，找到包含该位置的chunk
                for chunk in chunks_data['chunks']:
                    start_idx = chunk.get('start_index', 0)
                    end_idx = chunk.get('end_index', len(structured_content))
                    if start_idx <= item_position < end_idx:
                        return chunk.get('chunk_id')
                
                # 如果没找到，返回第一个chunk（作为默认值）
                if chunks_data['chunks']:
                    return chunks_data['chunks'][0].get('chunk_id')
                return None
            
            # 构建图片位置到structured_content索引的映射
            # 图片的position是段落索引，需要转换为structured_content索引
            image_position_to_structured_idx = {}
            structured_content = doc_data.get("structured_content", [])
            para_idx = 0
            for idx, item in enumerate(structured_content):
                if item.get("type") in ["paragraph", "heading"]:
                    # 检查这个段落是否有图片
                    images_in_item = item.get("images", [])
                    for img in images_in_item:
                        image_id = img.get("image_id")
                        if image_id:
                            image_position_to_structured_idx[image_id] = idx
                    para_idx += 1
            
            # 构建表格位置到structured_content索引的映射
            table_position_to_structured_idx = {}
            for idx, item in enumerate(structured_content):
                if item.get("type") == "table":
                    table_id = item.get("table_id")
                    if table_id:
                        table_position_to_structured_idx[table_id] = idx
            
            # 步骤7：创建图片Episode（关联到所属章节）
            image_episodes = []
            for idx, image in enumerate(doc_data.get("images", [])):
                # 构建图片Episode的内容（参考process_word_document的实现）
                image_desc = image.get("description", "图片")
                image_id = image.get("image_id", f"image_{idx+1}")
                section_title = image.get("section_title", "未知章节")
                relative_position = image.get("relative_position", 0.0)
                prev_context = image.get("prev_context", "")
                next_context = image.get("next_context", "")
                image_context = image.get("context", "")
                match_method = image.get("match_method", "未知")
                match_confidence = image.get("match_confidence", 0.0)
                
                # 构建图片URL
                image_url = f"/api/document-upload/{upload_id}/images/{image_id}"
                
                # 获取文件信息
                file_path = image.get("file_path", "")
                relative_path = image.get("relative_path", "")
                file_size = image.get("file_size", 0)
                file_format = image.get("file_format", "未知")
                
                # 如果file_size为0或file_format为"未知"，尝试从文件系统获取
                if file_size == 0 or file_format == "未知":
                    # 尝试从file_path获取
                    if file_path:
                        abs_file_path = file_path if os.path.isabs(file_path) else os.path.join("/app", file_path)
                        if os.path.exists(abs_file_path):
                            file_size = os.path.getsize(abs_file_path)
                            if file_format == "未知":
                                ext = os.path.splitext(abs_file_path)[1]
                                file_format = ext[1:].upper() if ext else "UNKNOWN"
                    # 如果还是0，尝试从relative_path构建路径
                    elif relative_path:
                        abs_file_path = os.path.join("/app", relative_path)
                        if os.path.exists(abs_file_path):
                            file_size = os.path.getsize(abs_file_path)
                            if file_format == "未知":
                                ext = os.path.splitext(abs_file_path)[1]
                                file_format = ext[1:].upper() if ext else "UNKNOWN"
                
                # 构建图片Episode的内容（增强版：包含更多元数据和上下文）
                image_content = f"""## 图片信息

**图片ID**: {image_id}
**描述**: {image_desc}
**文件路径**: {file_path}
**相对路径**: {relative_path}
**文件大小**: {file_size} 字节
**文件格式**: {file_format}
**匹配方法**: {match_method}
**匹配置信度**: {match_confidence:.2f}
**文档位置**: {relative_position:.1%}
"""
                
                # 添加章节信息
                if section_title:
                    image_content += f"**所属章节**: {section_title}\n\n"
                
                # 添加完整的上下文信息
                image_content += "### 上下文信息\n\n"
                if prev_context:
                    image_content += f"**前文**: {prev_context}\n\n"
                if image_context:
                    image_content += f"**当前段落**: {image_context}\n\n"
                if next_context:
                    image_content += f"**后文**: {next_context}\n\n"
                if not prev_context and not image_context and not next_context:
                    image_content += "无上下文信息\n\n"
                
                image_content += f"""### 图片链接
![{image_desc}]({image_url})

### 图片说明
这是一张从Word文档中提取的图片，位于文档的相应位置（位置: {relative_position:.1%}）。图片可能包含流程图、示意图、图表或其他可视化内容。

**匹配信息**: 通过{match_method}方法匹配，置信度为{match_confidence:.0%}。
"""
                
                # 确定图片所属的章节Episode
                image_id = image.get("image_id", "")
                structured_idx = image_position_to_structured_idx.get(image_id)
                section_episode_uuid_for_image = None
                
                if structured_idx is not None and chunks_data:
                    chunk_id = find_chunk_for_item(structured_idx, structured_content, chunks_data)
                    if chunk_id and chunk_id in section_episode_map:
                        section_episode_uuid_for_image = section_episode_map[chunk_id]
                        logger.info(f"图片 {image_id} 关联到章节Episode: {section_episode_uuid_for_image} (chunk_id: {chunk_id})")
                
                # 构建previous_episode_uuids：文档级Episode + 所属章节Episode（如果存在）
                previous_episode_uuids = [document_episode_uuid]
                if section_episode_uuid_for_image:
                    previous_episode_uuids.append(section_episode_uuid_for_image)
                
                image_episode = await graphiti.add_episode(
                    name=f"{document.file_name}_图片_{idx+1}_{image.get('image_id', '')}",
                    episode_body=image_content,
                    source_description="Word文档图片",
                    reference_time=doc_data["metadata"].get("created") or datetime.now(),
                    entity_types=ENTITY_TYPES,
                    edge_types=EDGE_TYPES,
                    edge_type_map=EDGE_TYPE_MAP,
                    group_id=group_id,
                    previous_episode_uuids=previous_episode_uuids
                )
                neo4j_client.execute_write(update_version_query, {
                    "episode_uuid": image_episode.episode.uuid,
                    "version": version,
                    "version_number": version_number,
                    "document_name": document.file_name,
                    "file_path": file_path_abs,
                    "original_filename": os.path.basename(file_path_abs)
                })
                image_episodes.append(image_episode.episode.uuid)
            
            # 步骤8：创建表格Episode（关联到所属章节）
            table_episodes = []
            for idx, table_data in enumerate(doc_data.get("tables", [])):
                # 格式化表格为标准Markdown格式（用于Episode内容）
                table_markdown = WordDocumentService._format_table_as_markdown(table_data)
                
                # 构建表格Episode的内容
                table_id = table_data.get('table_id', f'table_{idx+1}')
                table_content = f"""## 表格信息

**表格序号**: {idx+1}
**表格ID**: {table_id}
**行数**: {len(table_data.get('rows', []))}
**列数**: {len(table_data.get('headers', []))}

### 表格内容

{table_markdown}

### 表格说明
这是从Word文档中提取的表格数据，使用标准Markdown表格格式，包含结构化的信息。
"""
                
                # 确定表格所属的章节Episode
                structured_idx = table_position_to_structured_idx.get(table_id)
                section_episode_uuid_for_table = None
                
                if structured_idx is not None and chunks_data:
                    chunk_id = find_chunk_for_item(structured_idx, structured_content, chunks_data)
                    if chunk_id and chunk_id in section_episode_map:
                        section_episode_uuid_for_table = section_episode_map[chunk_id]
                        logger.info(f"表格 {table_id} 关联到章节Episode: {section_episode_uuid_for_table} (chunk_id: {chunk_id})")
                
                # 构建previous_episode_uuids：文档级Episode + 所属章节Episode（如果存在）
                previous_episode_uuids = [document_episode_uuid]
                if section_episode_uuid_for_table:
                    previous_episode_uuids.append(section_episode_uuid_for_table)
                
                table_episode = await graphiti.add_episode(
                    name=f"{document.file_name}_表格_{idx+1}_{table_id}",
                    episode_body=table_content,
                    source_description="Word文档表格",
                    reference_time=doc_data["metadata"].get("created") or datetime.now(),
                    entity_types=ENTITY_TYPES,
                    edge_types=EDGE_TYPES,
                    edge_type_map=EDGE_TYPE_MAP,
                    group_id=group_id,
                    previous_episode_uuids=previous_episode_uuids
                )
                neo4j_client.execute_write(update_version_query, {
                    "episode_uuid": table_episode.episode.uuid,
                    "version": version,
                    "version_number": version_number,
                    "document_name": document.file_name,
                    "file_path": file_path_abs,
                    "original_filename": os.path.basename(file_path_abs)
                })
                table_episodes.append(table_episode.episode.uuid)
            
            # 更新文档状态和document_id
            document.status = DocumentStatus.COMPLETED
            document.document_id = group_id
            db.commit()
            
            logger.info(f"文档处理完成: upload_id={upload_id}, group_id={group_id}")
            
            return DocumentProcessResponse(
                upload_id=upload_id,
                document_id=group_id,
                document_name=document.file_name,
                version=version,
                version_number=version_number,
                episodes={
                    "document": document_episode_uuid,
                    "sections": section_episodes,
                    "images": image_episodes,
                    "tables": table_episodes
                },
                statistics={
                    "total_episodes": 1 + len(section_episodes) + len(image_episodes) + len(table_episodes),
                    "total_sections": len(section_episodes),
                    "total_images": len(image_episodes),
                    "total_tables": len(table_episodes)
                }
            )
            
        except Exception as process_error:
            document.status = DocumentStatus.ERROR
            document.error_message = str(process_error)
            db.commit()
            logger.error(f"处理文档失败: {process_error}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"处理失败: {str(process_error)}")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"处理文档失败: {e}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")


@router.get("/{upload_id}/images/{image_id}")
async def get_uploaded_document_image(
    upload_id: int,
    image_id: str,
    db: Session = Depends(get_db)
):
    """
    获取上传文档中提取的图片
    """
    try:
        # 查询文档记录
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 使用 upload_{upload_id} 作为 document_id
        document_id = f"upload_{upload_id}"
        image_path = os.path.abspath(f"uploads/extracted_images/{document_id}/{image_id}.png")
        
        if not os.path.exists(image_path):
            # 尝试其他格式
            for ext in [".jpg", ".jpeg", ".gif", ".bmp"]:
                alt_path = os.path.abspath(f"uploads/extracted_images/{document_id}/{image_id}{ext}")
                if os.path.exists(alt_path):
                    image_path = alt_path
                    break
            else:
                raise HTTPException(status_code=404, detail="图片文件不存在")
        
        from fastapi.responses import FileResponse
        return FileResponse(
            path=image_path,
            media_type="image/png"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取图片失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@router.get("/{upload_id}/ole/{ole_id}")
async def get_uploaded_document_ole(
    upload_id: int,
    ole_id: str,
    view: str = Query("download", description="查看模式：download 或 preview"),
    db: Session = Depends(get_db)
):
    """
    获取上传文档中提取的嵌入文档（OLE对象）
    
    Args:
        view: "download" 下载文件, "preview" 在线预览（支持Excel和Word）
    """
    try:
        # 查询文档记录
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 使用 upload_{upload_id} 作为 document_id
        document_id = f"upload_{upload_id}"
        ole_dir = os.path.abspath(f"uploads/extracted_ole/{document_id}")
        
        # 查找文件（可能包含ole_前缀）
        ole_files = []
        if os.path.exists(ole_dir):
            import glob
            # 尝试多种可能的文件名格式
            patterns = [
                f"{ole_id}.*",
                f"ole_{ole_id}.*",
                f"{ole_id.replace('ole_', '')}.*"
            ]
            patterns = list(set(patterns))
            for pattern in patterns:
                ole_files.extend(glob.glob(os.path.join(ole_dir, pattern)))
        
        if not ole_files:
            raise HTTPException(status_code=404, detail="嵌入文档文件不存在")
        
        # 优先选择标准格式文件
        preferred_extensions = ['.xlsx', '.xls', '.docx', '.doc', '.pptx', '.ppt']
        ole_path = None
        
        for ext in preferred_extensions:
            for file_path in ole_files:
                if file_path.endswith(ext):
                    ole_path = file_path
                    break
            if ole_path:
                break
        
        if not ole_path:
            ole_path = ole_files[0]
        
        # 根据view参数决定返回方式
        if view == "preview":
            # 预览模式：返回HTML页面
            from fastapi.responses import HTMLResponse
            from docx import Document as DocxDocument
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            file_ext = os.path.splitext(ole_path)[1].lower()
            
            if file_ext in ['.xlsx', '.xls']:
                # Excel预览
                workbook = openpyxl.load_workbook(ole_path)
                sheet = workbook.active
                
                html_content = "<html><head><meta charset='UTF-8'><style>table{border-collapse:collapse;width:100%;}th,td{border:1px solid #ddd;padding:8px;text-align:left;}th{background-color:#f2f2f2;}</style></head><body><table>"
                
                for row in sheet.iter_rows(values_only=True):
                    html_content += "<tr>"
                    for cell in row:
                        html_content += f"<td>{cell or ''}</td>"
                    html_content += "</tr>"
                
                html_content += "</table></body></html>"
                return HTMLResponse(content=html_content)
            
            elif file_ext in ['.docx', '.doc']:
                # Word预览
                doc = DocxDocument(ole_path)
                html_content = "<html><head><meta charset='UTF-8'><style>body{font-family:'Microsoft YaHei',sans-serif;padding:20px;}</style></head><body>"
                
                for para in doc.paragraphs:
                    html_content += f"<p>{para.text}</p>"
                
                html_content += "</body></html>"
                return HTMLResponse(content=html_content)
            
            else:
                # 不支持预览，返回下载
                from fastapi.responses import FileResponse
                return FileResponse(path=ole_path)
        else:
            # 下载模式
            from fastapi.responses import FileResponse
            return FileResponse(path=ole_path)
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取嵌入文档失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@router.post("/{upload_id}/build-communities-async")
async def build_communities_async(
    upload_id: int,
    request: BuildCommunitiesRequest,
    provider: str = Query("local", description="LLM提供商"),
    use_thinking: bool = Query(False, description="是否启用Thinking模式（仅本地大模型支持）"),
    db: Session = Depends(get_db)
):
    """
    异步构建Community（当前文档或跨文档）
    
    参数：
    - scope: 'current' 或 'cross'
    - group_ids: 跨文档时提供的group_id列表（可选）
    - provider: LLM提供商
    - use_thinking: 是否启用Thinking模式（仅本地大模型支持）
    
    返回任务ID，可通过任务管理页面查看进度
    """
    try:
        import uuid
        from app.models.task_queue import TaskQueue, TaskStatus, TaskType
        
        # 生成任务ID
        celery_task_id = str(uuid.uuid4())
        
        # 查询文档
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 确定group_ids
        if request.scope == "current":
            if not document.document_id:
                raise HTTPException(status_code=400, detail="当前文档尚未完成处理，无法构建Community")
        elif request.scope == "cross":
            if not request.group_ids or len(request.group_ids) < 2:
                raise HTTPException(status_code=400, detail="跨文档构建需要至少选择2个文档")
        
        # 创建任务记录
        task = TaskQueue(
            task_id=celery_task_id,
            upload_id=upload_id,
            task_type=TaskType.BUILD_COMMUNITIES.value,
            status=TaskStatus.PENDING.value,
            progress=0,
            current_step="等待处理"
        )
        db.add(task)
        db.commit()
        db.refresh(task)
        
        # 提交Celery任务
        from app.tasks.build_communities import build_communities_task
        celery_task = build_communities_task.delay(
            task_id=celery_task_id,
            upload_id=upload_id,
            scope=request.scope,
            group_ids=request.group_ids,
            provider=provider,
            use_thinking=use_thinking if provider == "local" else False
        )
        
        # 更新任务ID为Celery任务ID
        if celery_task.id != celery_task_id:
            task.task_id = celery_task.id
            db.commit()
        
        return {
            "task_id": celery_task.id,
            "status": "pending",
            "message": "任务已提交，请前往任务管理页面查看进度"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"提交构建Community任务失败: {e}", exc_info=True)
        if task and task.id:
            db.delete(task)
            db.commit()
        raise HTTPException(status_code=500, detail=f"提交失败: {str(e)}")


@router.post("/{upload_id}/build-communities", response_model=BuildCommunitiesResponse, deprecated=True)
async def build_communities(
    upload_id: int,
    request: BuildCommunitiesRequest,
    provider: str = Query("qianwen", description="LLM提供商"),
    db: Session = Depends(get_db)
):
    """
    构建Community（当前文档或跨文档）
    
    参数：
    - scope: 'current' 或 'cross'
    - group_ids: 跨文档时提供的group_id列表（可选）
    """
    try:
        from app.core.graphiti_client import get_graphiti_instance
        from app.core.neo4j_client import neo4j_client
        from app.core.utils import serialize_neo4j_properties
        
        # 查询当前文档
        document = db.query(DocumentUpload).filter(DocumentUpload.id == upload_id).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 确定要使用的group_ids
        group_ids = []
        
        if request.scope == "current":
            # 当前文档
            if not document.document_id:
                raise HTTPException(status_code=400, detail="当前文档尚未完成处理，无法构建Community")
            group_ids = [document.document_id]
            logger.info(f"构建当前文档的Community: upload_id={upload_id}, group_id={document.document_id}")
        elif request.scope == "cross":
            # 跨文档
            if not request.group_ids or len(request.group_ids) < 2:
                raise HTTPException(status_code=400, detail="跨文档构建需要至少选择2个文档")
            group_ids = request.group_ids
            logger.info(f"构建跨文档的Community: upload_id={upload_id}, group_ids={group_ids}")
        else:
            raise HTTPException(status_code=400, detail=f"无效的构建范围: {request.scope}")
        
        # 获取Graphiti实例
        graphiti = get_graphiti_instance(provider)
        
        # 构建Community
        logger.info(f"开始构建Community，group_ids={group_ids}")
        communities_result = await graphiti.build_communities(group_ids=group_ids)
        logger.info(f"Graphiti build_communities 调用完成")
        
        # 直接从Neo4j查询Community节点（不依赖Graphiti的返回结果）
        # 因为Graphiti的返回结果结构可能不确定，但Community节点已经保存到Neo4j了
        # 构建查询条件：支持group_id是字符串或列表的情况
        if len(group_ids) == 1:
            # 单个group_id
            communities_query = """
            MATCH (c:Community)
            WHERE c.group_id = $group_id OR 
                  (c.group_id IS NOT NULL AND 
                   (toString(c.group_id) CONTAINS $group_id OR 
                    $group_id IN c.group_id))
            RETURN c.uuid as uuid, c.name as name, c.summary as summary, c.group_id as group_id
            ORDER BY c.name
            """
            communities_data = neo4j_client.execute_query(communities_query, {
                "group_id": group_ids[0]
            })
        else:
            # 多个group_id
            communities_query = """
            MATCH (c:Community)
            WHERE any(gid IN $group_ids WHERE 
              c.group_id = gid OR 
              (c.group_id IS NOT NULL AND 
               (toString(c.group_id) CONTAINS gid OR 
                gid IN c.group_id)))
            RETURN c.uuid as uuid, c.name as name, c.summary as summary, c.group_id as group_id
            ORDER BY c.name
            """
            communities_data = neo4j_client.execute_query(communities_query, {
                "group_ids": group_ids
            })
        
        logger.info(f"从Neo4j查询到 {len(communities_data)} 个Community节点")
        
        # ========== 翻译英文summary为中文 ==========
        logger.info(f"检查并翻译英文summary为中文")
        try:
            import re
            # 简单的英文检测：如果summary中超过50%是英文字符，认为是英文
            def is_mostly_english(text):
                if not text:
                    return False
                # 统计英文字母数量
                english_chars = len(re.findall(r'[a-zA-Z]', text))
                total_chars = len(re.findall(r'[a-zA-Z\u4e00-\u9fff]', text))  # 英文+中文
                if total_chars == 0:
                    return False
                return english_chars / total_chars > 0.5
            
            # 翻译英文summary
            for community_data in communities_data:
                comm_summary = community_data.get("summary", "")
                comm_name = community_data.get("name", "")
                community_uuid = community_data.get("uuid")
                
                # 检查name和summary是否为英文
                need_translate_name = is_mostly_english(comm_name) and not any('\u4e00' <= c <= '\u9fff' for c in comm_name)
                need_translate_summary = is_mostly_english(comm_summary) and not any('\u4e00' <= c <= '\u9fff' for c in comm_summary)
                
                if need_translate_name or need_translate_summary:
                    try:
                        # 构建翻译prompt
                        translate_prompt = "请将以下内容翻译成中文，保持原意不变：\n\n"
                        if need_translate_name:
                            translate_prompt += f"名称: {comm_name}\n"
                        if need_translate_summary:
                            translate_prompt += f"摘要: {comm_summary}\n"
                        translate_prompt += "\n请以JSON格式返回，格式：{\"name\": \"翻译后的名称\", \"summary\": \"翻译后的摘要\"}"
                        
                        # 调用LLM翻译（固定使用本地大模型）
                        from openai import AsyncOpenAI
                        from app.core.config import settings
                        
                        # 固定使用本地大模型进行翻译
                        local_base_url = settings.LOCAL_LLM_API_BASE_URL.rstrip('/')
                        if not local_base_url.endswith("/v1"):
                            if "/v1" not in local_base_url:
                                local_base_url = f"{local_base_url}/v1"
                        
                        openai_client = AsyncOpenAI(
                            api_key=settings.LOCAL_LLM_API_KEY,
                            base_url=local_base_url,
                            timeout=180.0
                        )
                        model = settings.LOCAL_LLM_MODEL
                        logger.info(f"使用本地大模型进行Community翻译: model={model}")
                        
                        # 调用OpenAI兼容接口进行翻译
                        llm_response = await openai_client.chat.completions.create(
                            model=model,
                            messages=[{"role": "user", "content": translate_prompt}],
                            temperature=0.3
                        )
                        translated_text = llm_response.choices[0].message.content.strip()
                        
                        # 解析JSON响应
                        import json
                        # 尝试提取JSON（可能包含markdown代码块）
                        # 先尝试直接解析
                        translated_data = None
                        try:
                            translated_data = json.loads(translated_text)
                        except:
                            # 如果直接解析失败，尝试提取JSON代码块
                            json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', translated_text, re.DOTALL)
                            if json_match:
                                try:
                                    translated_data = json.loads(json_match.group(1))
                                except:
                                    pass
                            # 如果还是失败，尝试提取第一个完整的JSON对象
                            if not translated_data:
                                json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', translated_text, re.DOTALL)
                                if json_match:
                                    try:
                                        translated_data = json.loads(json_match.group())
                                    except:
                                        pass
                        
                        if translated_data:
                            translated_name = translated_data.get("name", comm_name)
                            translated_summary = translated_data.get("summary", comm_summary)
                            
                            # 更新Neo4j中的Community节点
                            update_query = """
                            MATCH (c:Community {uuid: $uuid})
                            """
                            update_fields = []
                            params = {"uuid": community_uuid}
                            
                            if need_translate_name and translated_name != comm_name:
                                update_fields.append("c.name = $name")
                                params["name"] = translated_name
                                community_data["name"] = translated_name
                                logger.info(f"翻译Community名称: {comm_name} -> {translated_name}")
                            
                            if need_translate_summary and translated_summary != comm_summary:
                                update_fields.append("c.summary = $summary")
                                params["summary"] = translated_summary
                                community_data["summary"] = translated_summary
                                logger.info(f"翻译Community摘要: {comm_summary[:50]}... -> {translated_summary[:50]}...")
                            
                            if update_fields:
                                update_query += " SET " + ", ".join(update_fields)
                                update_query += " RETURN c.uuid as uuid"
                                neo4j_client.execute_write(update_query, params)
                                logger.info(f"Community {community_uuid} 翻译完成")
                    except Exception as translate_error:
                        logger.warning(f"翻译Community {community_uuid}失败: {translate_error}，保持原文")
                        continue
        except Exception as e:
            logger.warning(f"翻译Community summary时出错: {e}，继续处理")
        
        # ========== 为Community生成embedding（建议1） ==========
        logger.info(f"开始为 {len(communities_data)} 个Community生成embedding")
        try:
            # 生成查询文本的embedding（用于后续向量搜索）
            # 这里我们为每个Community生成name_embedding（基于name + summary）
            for community_data in communities_data:
                community_uuid = community_data.get("uuid")
                if not community_uuid:
                    continue
                
                # 检查是否已有embedding
                check_query = """
                MATCH (c:Community {uuid: $uuid})
                RETURN c.name_embedding IS NOT NULL as has_embedding
                """
                check_result = neo4j_client.execute_query(check_query, {"uuid": community_uuid})
                has_embedding = check_result[0].get("has_embedding", False) if check_result else False
                
                if has_embedding:
                    logger.debug(f"Community {community_uuid} 已有embedding，跳过")
                    continue
                
                # 构建要embed的文本（name + summary）
                comm_name = community_data.get("name", "")
                comm_summary = community_data.get("summary", "")
                text_to_embed = f"{comm_name} {comm_summary}".strip()
                
                if not text_to_embed:
                    logger.warning(f"Community {community_uuid} 没有name和summary，跳过embedding生成")
                    continue
                
                # 使用Graphiti的embedder生成embedding
                try:
                    embedding = await graphiti.embedder.create(text_to_embed)
                    
                    # 更新Community节点，添加name_embedding字段
                    update_query = """
                    MATCH (c:Community {uuid: $uuid})
                    SET c.name_embedding = $embedding
                    RETURN c.uuid as uuid
                    """
                    neo4j_client.execute_write(update_query, {
                        "uuid": community_uuid,
                        "embedding": embedding
                    })
                    logger.info(f"Community {community_uuid} embedding生成成功")
                except Exception as embed_error:
                    logger.warning(f"为Community {community_uuid}生成embedding失败: {embed_error}")
                    continue
            
            logger.info(f"Community embedding生成完成")
        except Exception as e:
            logger.warning(f"为Community生成embedding时出错: {e}，继续执行")
        
        # 处理返回结果
        communities = []
        total_entities = 0
        
        for community_data in communities_data:
            community_uuid = community_data.get("uuid")
            if not community_uuid:
                continue
                
            # 查询该Community包含的实体数量
            # Graphiti实际使用的关系类型是 HAS_MEMBER (Community -> Entity)
            # 同时支持其他可能的关系类型：CONTAINS, BELONGS_TO
            entity_count_query = """
            MATCH (c:Community {uuid: $community_uuid})
            OPTIONAL MATCH (c)-[:HAS_MEMBER|CONTAINS]->(e1:Entity)
            OPTIONAL MATCH (e2:Entity)-[:BELONGS_TO]->(c)
            RETURN count(DISTINCT e1) + count(DISTINCT e2) as entity_count
            """
            entity_result = neo4j_client.execute_query(entity_count_query, {
                "community_uuid": community_uuid
            })
            entity_count = entity_result[0].get("entity_count", 0) if entity_result else 0
            logger.info(f"Community {community_uuid} 包含 {entity_count} 个实体")
            total_entities += entity_count
            
            # 处理group_id
            community_group_ids = []
            group_id_value = community_data.get("group_id")
            if group_id_value:
                if isinstance(group_id_value, list):
                    community_group_ids = [str(gid) for gid in group_id_value]
                else:
                    community_group_ids = [str(group_id_value)]
            
            communities.append(CommunityInfo(
                uuid=str(community_uuid),
                name=community_data.get("name") or "未命名Community",
                summary=community_data.get("summary") or "",
                entity_count=entity_count,
                group_ids=community_group_ids
            ))
        
        logger.info(f"Community构建完成: 共 {len(communities)} 个Community，包含 {total_entities} 个实体")
        
        return BuildCommunitiesResponse(
            success=True,
            communities=communities,
            statistics={
                "total_communities": len(communities),
                "total_entities": total_entities
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"构建Community失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"构建失败: {str(e)}")

